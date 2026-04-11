import requests
import json
import os
import time

# 测试导出所有数据的API调用
def test_api_export():
    print("开始测试导出所有数据API")
    
    # 登录获取token
    login_url = 'http://localhost:5000/api/v1/auth/login'
    login_headers = {'Content-Type': 'application/json'}
    login_data = {
        'username': 'admin',
        'password': 'admin123'
    }
    
    print(f"登录请求: {login_url}")
    print(f"请求头: {login_headers}")
    print(f"请求数据: {login_data}")
    
    login_response = requests.post(login_url, headers=login_headers, json=login_data)
    print(f"登录响应状态码: {login_response.status_code}")
    print(f"登录响应内容: {login_response.text}")
    
    if login_response.status_code != 200:
        print("登录失败")
        return
    
    # 获取token
    login_result = login_response.json()
    if 'data' not in login_result or 'token' not in login_result['data']:
        print("登录响应格式错误")
        return
    
    token = login_result['data']['token']
    print(f"成功获取token: {token}")
    
    # 调用导出所有数据API
    export_url = 'http://localhost:5000/api/v1/export-all-data'
    export_headers = {
        'Authorization': f'Bearer {token}'
    }
    
    print(f"\n导出请求: {export_url}")
    print(f"请求头: {export_headers}")
    
    start_time = time.time()
    export_response = requests.get(export_url, headers=export_headers, stream=True)
    end_time = time.time()
    
    print(f"导出响应状态码: {export_response.status_code}")
    print(f"导出响应耗时: {end_time - start_time:.2f}秒")
    
    if export_response.status_code != 200:
        print(f"导出失败，响应内容: {export_response.text}")
        return
    
    # 保存响应内容到文件
    output_file = 'api_export_test.xlsx'
    with open(output_file, 'wb') as f:
        for chunk in export_response.iter_content(chunk_size=8192):
            f.write(chunk)
    
    file_size = os.path.getsize(output_file)
    print(f"成功保存导出文件: {output_file}")
    print(f"文件大小: {file_size} 字节")
    
    # 验证文件是否包含属性工作表
    from openpyxl import load_workbook
    try:
        wb = load_workbook(output_file)
        print(f"\n文件中包含的工作表: {wb.sheetnames}")
        
        # 检查是否有属性工作表
        has_properties_sheets = any(sheet.endswith('_属性') for sheet in wb.sheetnames)
        print(f"是否包含属性工作表: {has_properties_sheets}")
        
        # 检查第一个属性工作表的内容
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_属性'):
                print(f"\n检查属性工作表: {sheet_name}")
                ws = wb[sheet_name]
                
                # 打印前几行内容
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    print(f"行内容: {row}")
                
                # 检查行数
                rows_count = 0
                for row in ws:
                    rows_count += 1
                print(f"属性工作表行数: {rows_count}")
                break
    except Exception as e:
        print(f"验证文件失败: {str(e)}")

if __name__ == '__main__':
    test_api_export()
