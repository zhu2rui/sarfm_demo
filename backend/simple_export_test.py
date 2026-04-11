import requests
import json
import os
import time

# 简单的导出测试脚本
def simple_export_test():
    print("开始简单的导出测试")
    
    # 1. 登录获取token
    print("\n1. 登录获取token...")
    login_url = 'http://localhost:5000/api/v1/auth/login'
    login_headers = {
        'Content-Type': 'application/json'
    }
    login_data = {
        'username': 'admin',
        'password': 'admin123'
    }
    
    login_response = requests.post(login_url, headers=login_headers, json=login_data)
    print(f"登录响应状态码: {login_response.status_code}")
    
    if login_response.status_code != 200:
        print(f"登录失败，响应内容: {login_response.text}")
        return
    
    login_result = login_response.json()
    token = login_result['data']['token']
    print(f"成功获取token: {token}")
    
    # 2. 调用导出所有数据API
    print("\n2. 调用导出所有数据API...")
    export_url = 'http://localhost:5000/api/v1/export-all-data'
    export_headers = {
        'Authorization': f'Bearer {token}'
    }
    
    start_time = time.time()
    export_response = requests.get(export_url, headers=export_headers, stream=True)
    end_time = time.time()
    
    print(f"导出响应状态码: {export_response.status_code}")
    print(f"导出耗时: {end_time - start_time:.2f}秒")
    print(f"响应头: {dict(export_response.headers)}")
    
    if export_response.status_code == 200:
        # 保存响应内容到文件
        output_file = 'simple_export_test.xlsx'
        with open(output_file, 'wb') as f:
            for chunk in export_response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        file_size = os.path.getsize(output_file)
        print(f"成功保存导出文件: {output_file}")
        print(f"文件大小: {file_size} 字节")
        
        # 验证文件内容
        from openpyxl import load_workbook
        try:
            wb = load_workbook(output_file)
            print(f"\n3. 验证文件内容...")
            print(f"文件中包含的工作表: {wb.sheetnames}")
            
            # 检查是否有属性工作表
            has_properties_sheets = any(sheet.endswith('_属性') for sheet in wb.sheetnames)
            print(f"是否包含属性工作表: {has_properties_sheets}")
            
            if has_properties_sheets:
                # 显示第一个属性工作表的内容
                properties_sheet = next(sheet for sheet in wb.sheetnames if sheet.endswith('_属性'))
                ws = wb[properties_sheet]
                print(f"\n{properties_sheet}工作表内容:")
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                    print(f"第{i}行: {row}")
        except Exception as e:
            print(f"验证文件失败: {str(e)}")
    else:
        print(f"导出失败，响应内容: {export_response.text}")

if __name__ == '__main__':
    simple_export_test()
