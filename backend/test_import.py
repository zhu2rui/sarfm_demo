import os
import sys
import json
from io import BytesIO
from openpyxl import load_workbook

# 添加当前目录到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# 导入Flask应用和模型
from app import app, db, TableStructure, InventoryData, AutoIncrementSequence

# 测试导入数据的逻辑
def test_import_logic():
    with app.app_context():
        print("开始测试导入数据功能")
        
        # 读取测试导出的文件
        test_file = 'test_export_direct.xlsx'
        if not os.path.exists(test_file):
            print(f"测试文件 {test_file} 不存在，请先运行导出测试")
            return
        
        print(f"成功读取测试文件: {test_file}")
        
        # 加载工作簿
        wb = load_workbook(test_file)
        
        # 首先区分数据工作表和属性工作表
        data_sheets = []
        properties_sheets = {}
        
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_属性'):
                # 属性工作表，提取原始表格名称
                original_table_name = sheet_name[:-3]  # 移除'_属性'后缀
                properties_sheets[original_table_name] = sheet_name
            else:
                # 数据工作表
                data_sheets.append(sheet_name)
        
        print(f"识别到数据工作表: {data_sheets}")
        print(f"识别到属性工作表: {properties_sheets}")
        
        # 处理每个数据工作表
        for sheet_name in data_sheets:
            print(f"\n处理数据工作表: {sheet_name}")
            ws = wb[sheet_name]
            
            # 获取表头
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            print(f"表头: {header_row}")
            
            # 检查是否包含创建时间列
            created_at_column_index = None
            for i, header in enumerate(header_row):
                if header == '创建时间':
                    created_at_column_index = i
                    break
            
            if created_at_column_index is None:
                print(f"缺少创建时间列")
                continue
            
            # 获取数据列名
            data_headers = header_row[:created_at_column_index]
            print(f"数据列名: {data_headers}")
            
            # 查找对应的属性工作表
            columns_def = []
            
            if sheet_name in properties_sheets:
                # 找到属性工作表
                properties_sheet_name = properties_sheets[sheet_name]
                properties_ws = wb[properties_sheet_name]
                print(f"找到对应的属性工作表: {properties_sheet_name}")
                
                # 读取属性工作表数据
                properties_rows = list(properties_ws.iter_rows(min_row=1, values_only=True))
                
                if len(properties_rows) < 2:  # 至少需要表头和一行数据
                    print(f"属性工作表格式不正确，只有{len(properties_rows)}行数据")
                else:
                    # 解析属性工作表
                    properties_header = properties_rows[0]
                    
                    # 验证属性表头
                    expected_headers = ['column_name', 'data_type', 'dropDown', 'autoIncrement', 'prefix', 'hidden']
                    if list(properties_header) == expected_headers:
                        print("属性表头格式正确")
                        
                        # 使用属性工作表定义的列配置
                        for row in properties_rows[1:]:
                            if len(row) >= 6:
                                column_def = {
                                    'column_name': row[0] or '',
                                    'data_type': row[1] or 'string',
                                    'dropDown': bool(row[2]),
                                    'autoIncrement': bool(row[3]),
                                    'prefix': row[4] or '',
                                    'hidden': bool(row[5])
                                }
                                columns_def.append(column_def)
                                print(f"读取到列属性: {column_def}")
                    else:
                        print(f"属性表头格式不正确，预期: {expected_headers}，实际: {list(properties_header)}")
            
            if columns_def:
                print(f"成功读取{len(columns_def)}列属性")
            else:
                print("没有读取到列属性，使用默认配置")
                # 使用默认配置
                for header in data_headers:
                    columns_def.append({
                        'column_name': header,
                        'data_type': 'string',
                        'hidden': False
                    })
            
            # 打印最终的列配置
            print(f"最终列配置: {json.dumps(columns_def, indent=2, ensure_ascii=False)}")
            
            # 这里可以添加实际的导入逻辑测试，比如创建表格结构等
            # 为了简单起见，我们只打印列配置，不实际创建表格
            
        print("\n导入测试完成")

if __name__ == '__main__':
    test_import_logic()
