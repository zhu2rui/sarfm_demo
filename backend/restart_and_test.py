import os
import sys
import time

# 重启Flask开发服务器并测试导出功能
def restart_and_test():
    print("开始重启Flask开发服务器并测试导出功能")
    
    # 1. 杀死可能正在运行的Flask进程
    print("\n1. 杀死可能正在运行的Flask进程...")
    try:
        # 使用taskkill命令杀死Flask进程
        os.system('taskkill /f /im python.exe /fi "WINDOWTITLE eq Python*app.py*"')
        os.system('taskkill /f /im python.exe /fi "WINDOWTITLE eq *Flask*"')
        print("成功杀死Flask进程")
    except Exception as e:
        print(f"杀死Flask进程失败: {str(e)}")
    
    time.sleep(2)  # 等待2秒，确保进程完全终止
    
    # 2. 启动Flask开发服务器
    print("\n2. 启动Flask开发服务器...")
    # 使用start命令在新窗口启动Flask服务器
    os.system('start "Flask Server" python app.py')
    print("Flask服务器已在新窗口启动")
    
    time.sleep(5)  # 等待5秒，确保服务器完全启动
    
    # 3. 运行导出测试脚本
    print("\n3. 运行导出测试脚本...")
    try:
        # 运行导出测试脚本
        os.system('python test_export_direct.py')
        print("导出测试脚本运行完成")
        
        # 检查生成的文件
        if os.path.exists('test_export_direct.xlsx'):
            print("\n4. 检查生成的文件...")
            from openpyxl import load_workbook
            wb = load_workbook('test_export_direct.xlsx')
            print(f"文件中包含的工作表: {wb.sheetnames}")
            
            # 检查是否有属性工作表
            has_properties_sheets = any(sheet.endswith('_属性') for sheet in wb.sheetnames)
            print(f"是否包含属性工作表: {has_properties_sheets}")
            
            if has_properties_sheets:
                print("\n导出功能测试成功！生成的Excel文件包含了属性工作表。")
            else:
                print("\n导出功能测试失败！生成的Excel文件没有包含属性工作表。")
        else:
            print("\n导出测试脚本没有生成测试文件。")
    except Exception as e:
        print(f"运行导出测试脚本失败: {str(e)}")

if __name__ == '__main__':
    restart_and_test()
