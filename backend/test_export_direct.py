import os
import sys
import json
from io import BytesIO
from openpyxl import Workbook

# 添加当前目录到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# 导入Flask应用和模型
from app import app, db, TableStructure, InventoryData, User

# 测试导出所有数据的逻辑
def test_export_logic():
    with app.app_context():
        print("开始执行导出所有数据功能")
        
        # 创建工作簿
        wb = Workbook()
        print("成功创建工作簿")
        
        # 获取所有表格结构
        tables = TableStructure.query.all()
        print(f"成功获取{len(tables)}个表格结构")
        
        if not tables:
            print("没有可导出的表格数据")
            return
        
        # 删除默认工作表
        for sheet in wb.sheetnames:
            if sheet in ['Sheet', 'Sheet1']:
                del wb[sheet]
                print(f"成功删除默认工作表: {sheet}")
                break
        
        # 为每个表格创建工作表
        for table in tables:
            table_name = table.table_name
            print(f"处理表格: {table_name} (ID: {table.id})")
            
            # 解析columns
            try:
                columns = json.loads(table.columns)
                print(f"成功解析表格{table_name}的columns: {len(columns)}列")
                print(f"列信息: {columns}")
            except json.JSONDecodeError as e:
                print(f"解析表格{table_name}的columns失败: {str(e)}")
                print(f"columns原始内容: {table.columns}")
                continue
            
            # 创建数据工作表
            data_sheet_name = table_name[:31]
            data_ws = wb.create_sheet(title=data_sheet_name)
            print(f"成功创建数据工作表: {data_sheet_name}")
            
            # 创建属性工作表
            properties_sheet_name = f"{table_name[:25]}_属性"
            properties_ws = wb.create_sheet(title=properties_sheet_name)
            print(f"成功创建列属性工作表: {properties_sheet_name}")
            
            # 写入列属性
            properties_headers = ['column_name', 'data_type', 'dropDown', 'autoIncrement', 'prefix', 'hidden']
            properties_ws.append(properties_headers)
            
            for col in columns:
                properties_row = [
                    col.get('column_name', ''),
                    col.get('data_type', 'string'),
                    col.get('dropDown', False),
                    col.get('autoIncrement', False),
                    col.get('prefix', ''),
                    col.get('hidden', False)
                ]
                properties_ws.append(properties_row)
                print(f"写入列属性: {properties_row}")
            
            # 获取表格数据
            all_data = InventoryData.query.filter_by(table_id=table.id).all()
            print(f"成功获取表格{table_name}的{len(all_data)}条数据")
            
            # 生成数据工作表表头
            data_headers = [col['column_name'] for col in columns] + ['创建时间']
            data_ws.append(data_headers)
            print(f"成功写入数据工作表表头: {data_headers}")
            
            # 写数据行
            for i, item in enumerate(all_data):
                try:
                    item_data = json.loads(item.data)
                    row = []
                    for col in columns:
                        value = item_data.get(col['column_name'], '')
                        if value is None:
                            value = ''
                        if isinstance(value, dict):
                            if '_text' in value:
                                value = value['_text']
                            else:
                                value = json.dumps(value)
                        if not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                        row.append(value)
                    row.append(item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    data_ws.append(row)
                    print(f"成功写入表格{table_name}的数据行 {i+1}")
                except Exception as e:
                    print(f"处理表格{table_name}的数据行 {i+1}失败: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            print(f"成功写入表格{table_name}的所有数据")
        
        # 保存工作簿
        output_file = 'test_export_direct.xlsx'
        wb.save(output_file)
        print(f"成功保存工作簿到文件: {output_file}")
        
        # 检查文件是否存在
        if os.path.exists(output_file):
            print(f"文件已成功创建，大小: {os.path.getsize(output_file)} 字节")
        else:
            print("文件创建失败")
        
        # 检查工作簿中的工作表
        print(f"工作簿中包含的工作表: {wb.sheetnames}")

if __name__ == '__main__':
    test_export_logic()
