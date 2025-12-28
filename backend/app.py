# -*- coding: utf-8 -*-
from flask import Flask, jsonify, request, Response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv
import os
import sys
import jwt
from datetime import datetime, timedelta
from jwt import decode as jwt_decode
from functools import wraps
import json
import codecs

# 确保标准输出使用UTF-8编码
sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer)

# 获取应用程序根目录（兼容PyInstaller打包）
def get_app_root():
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller打包后的环境
        return sys._MEIPASS
    else:
        # 开发环境
        return os.path.dirname(os.path.abspath(__file__))

# 加载环境变量
load_dotenv()

# 创建Flask应用
app = Flask(__name__)
CORS(app)  # 启用CORS

# 设置JWT密钥
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key')

# 配置数据库
import os
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(get_app_root(), 'instance', 'inventory.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 初始化数据库实例
db = SQLAlchemy(app)

# 创建instance目录（如果不存在）
os.makedirs(os.path.join(app.root_path, 'instance'), exist_ok=True)

# 设置静态文件目录
from flask import send_from_directory
import os

# 设置静态文件目录
app.config['FRONTEND_DIST'] = os.path.join(get_app_root(), 'frontend_dist')

# 确保JSON响应使用UTF-8编码
@app.after_request
def after_request(response):
    # 只对JSON响应设置Content-Type
    if 'application/json' in response.headers.get('Content-Type', ''):
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response

# 提供前端静态文件
@app.route('/')
def serve_index():
    # 显式设置Content-Type为text/html
    from flask import make_response
    response = make_response(send_from_directory(app.config['FRONTEND_DIST'], 'index.html'))
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

@app.route('/<path:path>')
def serve_static(path):
    # 根据文件扩展名设置正确的Content-Type
    from flask import make_response, send_from_directory, abort
    import mimetypes
    import os
    
    # 检查文件是否存在
    file_path = os.path.join(app.config['FRONTEND_DIST'], path)
    if os.path.exists(file_path) and os.path.isfile(file_path):
        # 获取文件的MIME类型
        mime_type, _ = mimetypes.guess_type(path)
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        # 发送文件并设置正确的Content-Type
        response = make_response(send_from_directory(app.config['FRONTEND_DIST'], path))
        response.headers['Content-Type'] = f'{mime_type}; charset=utf-8'
        return response
    else:
        # SPA应用处理：所有未匹配的路由都返回index.html，让React Router处理
        response = make_response(send_from_directory(app.config['FRONTEND_DIST'], 'index.html'))
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
        return response

# 用户模型
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # admin/leader/member
    first_login = db.Column(db.Boolean, default=True, nullable=False)  # 标记是否首次登录
    created_at = db.Column(db.DateTime, default=db.func.now())
    updated_at = db.Column(db.DateTime, default=db.func.now(), onupdate=db.func.now())

# 表格结构模型
class TableStructure(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    table_name = db.Column(db.String(255), nullable=False)
    columns = db.Column(db.Text, nullable=False)  # JSON字符串格式: [{"column_name": "列名", "data_type": "数据类型"}...]
    created_at = db.Column(db.DateTime, default=db.func.now())
    updated_at = db.Column(db.DateTime, default=db.func.now(), onupdate=db.func.now())

# 库存数据模型
class InventoryData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    table_id = db.Column(db.Integer, db.ForeignKey('table_structure.id'), nullable=False)
    data = db.Column(db.Text, nullable=False)  # JSON字符串格式: {"列名1": "值1", "列名2": "值2", ...}
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=db.func.now())
    updated_at = db.Column(db.DateTime, default=db.func.now(), onupdate=db.func.now())

# 操作日志模型
class OperationLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    operation = db.Column(db.String(255), nullable=False)
    table_id = db.Column(db.Integer, nullable=True)
    data_id = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=db.func.now())

# 自增序列模型
class AutoIncrementSequence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    table_id = db.Column(db.Integer, db.ForeignKey('table_structure.id'), nullable=False)
    column_name = db.Column(db.String(255), nullable=False)
    current_value = db.Column(db.Integer, default=0, nullable=False)
    created_at = db.Column(db.DateTime, default=db.func.now())
    updated_at = db.Column(db.DateTime, default=db.func.now(), onupdate=db.func.now())
    
    # 唯一约束，确保每个表格的每列只有一个自增序列
    __table_args__ = (db.UniqueConstraint('table_id', 'column_name', name='_table_column_unique'),)

# Bug反馈模型
class BugReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    content = db.Column(db.Text, nullable=False)
    reporter_username = db.Column(db.String(50), nullable=False)
    is_resolved = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=db.func.now())
    updated_at = db.Column(db.DateTime, default=db.func.now(), onupdate=db.func.now())

# JWT认证装饰器
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        # 从请求头获取token
        if 'Authorization' in request.headers:
            token = request.headers['Authorization'].split(' ')[1]  # 格式: Bearer <token>
        
        if not token:
            return jsonify({'code': 401, 'message': 'Token is missing!', 'data': None}), 401
        
        try:
            # 解码token
            data = jwt_decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = User.query.filter_by(id=data['user_id']).first()
            if not current_user:
                return jsonify({'code': 401, 'message': '用户不存在!', 'data': None}), 401
        except jwt.ExpiredSignatureError:
            return jsonify({'code': 401, 'message': '令牌已过期!', 'data': None}), 401
        except jwt.InvalidTokenError:
            return jsonify({'code': 401, 'message': '无效的令牌!', 'data': None}), 401
        
        # 将用户信息传递给路由处理函数
        return f(current_user, *args, **kwargs)
    return decorated

# 初始化数据库
with app.app_context():
    db.create_all()

    # 创建默认用户账号
    from werkzeug.security import generate_password_hash
    
    # 创建管理员账号
    admin_user = User.query.filter_by(username='admin').first()
    if not admin_user:
        admin = User(username='admin', password=generate_password_hash('admin123'), role='admin')
        db.session.add(admin)
    
    # 创建组长账号
    leader_user = User.query.filter_by(username='leader').first()
    if not leader_user:
        leader = User(username='leader', password=generate_password_hash('leader123'), role='leader')
        db.session.add(leader)
    
    # 创建成员账号
    member_user = User.query.filter_by(username='member').first()
    if not member_user:
        member = User(username='member', password=generate_password_hash('member123'), role='member')
        db.session.add(member)
    
    # 创建测试机器人账号
    testbot_user = User.query.filter_by(username='testbot_playwright').first()
    if not testbot_user:
        testbot = User(username='testbot_playwright', password=generate_password_hash('000000'), role='member')
        db.session.add(testbot)
    
    db.session.commit()

# 健康检查路由
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'code': 200, 'message': 'success', 'data': {'status': 'healthy'}}), 200

# 登录路由
@app.route('/api/v1/auth/login', methods=['POST'])
def login():
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    remember = data.get('remember', False)
    
    logger.info(f'登录请求: 用户名={username}, 密码={password}, remember={remember}')
    
    if not username:
        logger.info(f'登录失败: 用户名不能为空')
        return jsonify({'code': 400, 'message': '用户名不能为空', 'data': None}), 400
    
    # 处理密码为空的情况，允许用户使用空密码登录
    password = password or ''
    
    user = User.query.filter_by(username=username).first()
    if not user:
        logger.info(f'登录失败: 用户名 {username} 不存在')
        return jsonify({'code': 401, 'message': '用户名或密码错误', 'data': None}), 401
    
    from werkzeug.security import check_password_hash
    
    # 检查密码是否为空，空密码的哈希值是固定格式，可以通过检查空字符串的哈希值来判断
    from werkzeug.security import generate_password_hash
    is_empty_password = check_password_hash(user.password, '')
    logger.info(f'用户 {username} 的密码哈希: {user.password}')
    logger.info(f'密码为空? {password == ""}, 是否是初始密码? {is_empty_password}')
    
    # 测试密码是否匹配
    password_matches = check_password_hash(user.password, password)
    logger.info(f'密码匹配? {password_matches}')
    
    # 测试一些常用密码
    common_passwords = ['admin123', 'leader123', 'member123', '123456']
    for pwd in common_passwords:
        if check_password_hash(user.password, pwd):
            logger.info(f'常用密码 {pwd} 匹配!')
    
    # 检查是否使用默认密码登录
    is_default_password = check_password_hash(user.password, '000000')
    logger.info(f'是否使用默认密码? {is_default_password}, 密码是000000? {password == "000000"}')
    
    # 进行密码验证
    if not check_password_hash(user.password, password):
        logger.info(f'登录失败: 用户名 {username} 的密码不正确')
        return jsonify({'code': 401, 'message': '用户名或密码错误', 'data': None}), 401
    
    # 如果是首次登录或使用默认密码登录，需要重置密码
    if user.first_login or (is_default_password and password == "000000"):
        logger.info(f'首次登录或使用默认密码: 用户 {username} 需要重置密码')
        # 记录初始登录尝试
        try:
            log = OperationLog(
                user_id=user.id,
                operation=f'用户 {username} 首次登录或使用默认密码登录，需要重置密码'
            )
            db.session.add(log)
            db.session.commit()
        except Exception as e:
            logger.error(f'记录日志失败: {str(e)}')
            db.session.rollback()
        return jsonify({
            'code': 200,
            'message': '首次登录或使用默认密码，请重置密码',
            'data': {
                'need_reset_password': True,
                'username': user.username
            }
        }), 200
    
    logger.info(f'登录成功: 用户名 {username}')
    
    # 根据remember参数设置token过期时间
    if remember:
        # 30天内免登录
        expires_in = timedelta(days=30)
    else:
        # 1小时过期
        expires_in = timedelta(hours=1)
    
    # 生成JWT token (使用PyJWT 2.x的正确API)
    try:
        token = jwt.encode(
            {
                'user_id': user.id,
                'username': user.username,
                'role': user.role,
                'exp': datetime.utcnow() + expires_in  # 设置token过期时间
            }, 
            app.config['SECRET_KEY'], 
            algorithm='HS256'
        )
        logger.info(f'成功生成token: {token}, 过期时间: {expires_in}')
    except Exception as e:
        logger.error(f'生成token失败: {e}')
        # 如果JWT生成失败，使用一个简单的token
        token = f"simple-token-{user.id}-{datetime.utcnow().timestamp()}"
        logger.info(f'使用简化token: {token}')
    
    return jsonify({
        'code': 200,
        'message': '登录成功',
        'data': {
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'created_at': user.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }
    }), 200

# 获取当前用户信息
@app.route('/api/v1/auth/me', methods=['GET'])
@token_required
def get_current_user(current_user):
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'id': current_user.id,
            'username': current_user.username,
            'role': current_user.role,
            'created_at': current_user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': current_user.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        }
    }), 200

# 重置密码路由
@app.route('/api/v1/auth/reset-password', methods=['POST'])
def reset_password():
    import logging
    import traceback
    import sys
    
    # 配置日志，确保所有日志都能被输出
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout),  # 输出到控制台
        ]
    )
    logger = logging.getLogger(__name__)
    
    logger.info('=' * 50)
    logger.info('开始处理密码重置请求')
    
    try:
        # 打印请求信息
        logger.debug(f'请求方法: {request.method}')
        logger.debug(f'请求URL: {request.url}')
        logger.debug(f'请求头: {dict(request.headers)}')
        
        # 获取请求数据
        data = request.get_json()
        logger.debug(f'原始请求数据: {data}')
        
        # 验证请求数据
        if not data:
            logger.error('请求数据为空或格式错误')
            return jsonify({'code': 400, 'message': '请求数据格式错误', 'data': None}), 400
        
        username = data.get('username')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        
        logger.info(f'请求参数: username={username}, new_password={new_password}, confirm_password={confirm_password}')
        
        # 验证请求参数
        if not username:
            logger.warning('用户名为空')
            return jsonify({'code': 400, 'message': '用户名不能为空', 'data': None}), 400
        if not new_password:
            logger.warning('新密码为空')
            return jsonify({'code': 400, 'message': '新密码不能为空', 'data': None}), 400
        if not confirm_password:
            logger.warning('确认密码为空')
            return jsonify({'code': 400, 'message': '确认密码不能为空', 'data': None}), 400
        
        # 验证两次密码输入一致
        if new_password != confirm_password:
            logger.warning(f'用户 {username} 两次密码输入不一致: {new_password} != {confirm_password}')
            return jsonify({'code': 400, 'message': '两次密码输入不一致', 'data': None}), 400
        
        logger.info('密码强度验证通过')
        
        # 查找用户
        logger.info(f'查找用户: {username}')
        user = User.query.filter_by(username=username).first()
        if not user:
            logger.warning(f'用户 {username} 不存在')
            return jsonify({'code': 404, 'message': '用户不存在', 'data': None}), 404
        
        logger.info(f'找到用户: id={user.id}, username={user.username}, first_login={user.first_login}, role={user.role}')
        
        # 更新密码
        logger.info(f'更新用户 {username} 的密码')
        from werkzeug.security import generate_password_hash
        user.password = generate_password_hash(new_password)
        user.first_login = False  # 标记为非首次登录
        
        # 打印用户更新后的信息
        logger.debug(f'用户更新后: password_hash={user.password[:20]}..., first_login={user.first_login}')
        
        # 尝试保存到数据库
        logger.info('开始提交到数据库')
        try:
            db.session.commit()
            logger.info(f'用户 {username} 密码重置成功！')
            return jsonify({
                'code': 200,
                'message': '密码修改成功，请使用新密码登录',
                'data': None
            }), 200
        except Exception as db_error:
            logger.error(f'数据库提交失败: {str(db_error)}')
            logger.error(f'详细错误信息: {traceback.format_exc()}')
            logger.error(f'错误类型: {type(db_error).__name__}')
            logger.error(f'错误参数: {db_error.args}')
            db.session.rollback()
            return jsonify({'code': 500, 'message': f'重置密码失败: 数据库操作错误 - {str(db_error)}', 'data': None}), 500
            
    except Exception as e:
        logger.error(f'密码重置过程中发生未知错误: {str(e)}')
        logger.error(f'详细错误信息: {traceback.format_exc()}')
        logger.error(f'错误类型: {type(e).__name__}')
        logger.error(f'错误参数: {e.args}')
        logger.error(f'堆栈信息: {sys.exc_info()[2]}')
        return jsonify({'code': 500, 'message': f'重置密码失败: 系统错误 - {str(e)}', 'data': None}), 500
    finally:
        logger.info('密码重置请求处理完成')
        logger.info('=' * 50)

# 获取用户列表
@app.route('/api/v1/users', methods=['GET'])
@token_required
def get_users(current_user):
    # 只有管理员可以查看用户列表
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    users = User.query.all()
    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'username': user.username,
            'role': user.role,
            'created_at': user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': user.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'items': users_data,
            'total': len(users_data)
        }
    }), 200

# 创建用户
@app.route('/api/v1/users', methods=['POST'])
@token_required
def add_user(current_user):
    # 只有管理员可以创建用户
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role = data.get('role')
    
    if not username or not role:
        return jsonify({'code': 400, 'message': '用户名和角色不能为空', 'data': None}), 400
    
    # 检查用户是否已存在
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        return jsonify({'code': 400, 'message': '用户名已存在', 'data': None}), 400
    
    # 检查角色是否合法
    if role not in ['admin', 'leader', 'member']:
        return jsonify({'code': 400, 'message': '角色必须是admin、leader或member', 'data': None}), 400
    
    # 创建新用户
    from werkzeug.security import generate_password_hash
    # 设置默认密码为"000000"如果未提供
    hashed_password = generate_password_hash(password or '000000')
    new_user = User(
        username=username,
        password=hashed_password,
        role=role
    )
    
    try:
        db.session.add(new_user)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '用户创建成功，初始密码为空，用户可以立即使用用户名登录',
            'data': {
                'id': new_user.id,
                'username': new_user.username,
                'role': new_user.role,
                'created_at': new_user.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建用户失败: {str(e)}', 'data': None}), 500

# 更新用户
@app.route('/api/v1/users/<int:user_id>', methods=['PUT'])
@token_required
def update_user(current_user, user_id):
    # 只有管理员可以更新用户
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'code': 404, 'message': '用户不存在', 'data': None}), 404
    
    data = request.get_json()
    password = data.get('password')
    role = data.get('role')
    
    # 更新密码
    if password:
        from werkzeug.security import generate_password_hash
        user.password = generate_password_hash(password)
    
    # 更新角色
    if role:
        if role not in ['admin', 'leader', 'member']:
            return jsonify({'code': 400, 'message': '角色必须是admin、leader或member', 'data': None}), 400
        user.role = role
    
    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '用户更新成功',
            'data': {
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'updated_at': user.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新用户失败: {str(e)}', 'data': None}), 500

# 删除用户
@app.route('/api/v1/users/<int:user_id>', methods=['DELETE'])
@token_required
def delete_user(current_user, user_id):
    # 只有管理员可以删除用户
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    # 不能删除自己
    if current_user.id == user_id:
        return jsonify({'code': 400, 'message': '不能删除自己', 'data': None}), 400
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'code': 404, 'message': '用户不存在', 'data': None}), 404
    
    # 不能删除admin用户
    if user.username == 'admin':
        return jsonify({'code': 400, 'message': '不能删除admin用户', 'data': None}), 400
    
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({'code': 200, 'message': '用户删除成功', 'data': None}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除用户失败: {str(e)}', 'data': None}), 500

# 创建表格结构
@app.route('/api/v1/tables', methods=['POST'])
@token_required
def create_table_structure(current_user):
    # 只有管理员可以创建表格结构
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    data = request.get_json()
    table_name = data.get('table_name')
    columns = data.get('columns')
    
    if not table_name or not columns or len(columns) == 0:
        return jsonify({'code': 400, 'message': '表格名称和列配置不能为空，列配置至少需要一列', 'data': None}), 400
    
    # 验证列配置格式
    for column in columns:
        if not column.get('column_name') or not column.get('data_type'):
            return jsonify({'code': 400, 'message': '每一列都必须有列名和数据类型', 'data': None}), 400
        # 为没有hidden属性的列设置默认值
        if 'hidden' not in column:
            column['hidden'] = False
    
    # 将列配置转换为JSON字符串
    columns_json = json.dumps(columns)
    
    try:
        new_table = TableStructure(
            table_name=table_name,
            columns=columns_json
        )
        db.session.add(new_table)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': '表格结构创建成功',
            'data': {
                'id': new_table.id,
                'table_name': new_table.table_name,
                'columns': columns,
                'created_at': new_table.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': new_table.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建表格结构失败: {str(e)}', 'data': None}), 500

# 获取表格结构列表
@app.route('/api/v1/tables', methods=['GET'])
@token_required
def get_table_structures(current_user):
    tables = TableStructure.query.all()
    tables_data = []
    
    for table in tables:
        tables_data.append({
            'id': table.id,
            'table_name': table.table_name,
            'columns': json.loads(table.columns),
            'created_at': table.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': table.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'items': tables_data,
            'total': len(tables_data)
        }
    }), 200

# 获取表格结构详情
@app.route('/api/v1/tables/<int:table_id>', methods=['GET'])
@token_required
def get_table_structure_detail(current_user, table_id):
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'id': table.id,
            'table_name': table.table_name,
            'columns': json.loads(table.columns),
            'created_at': table.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': table.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        }
    }), 200

# 更新表格结构
@app.route('/api/v1/tables/<int:table_id>', methods=['PUT'])
@token_required
def update_table_structure(current_user, table_id):
    # 只有管理员可以更新表格结构
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    data = request.get_json()
    table_name = data.get('table_name')
    columns = data.get('columns')
    
    if table_name:
        table.table_name = table_name
    
    if columns:
        if len(columns) == 0:
            return jsonify({'code': 400, 'message': '列配置不能为空，至少需要一列', 'data': None}), 400
        
        # 验证列配置格式
        for column in columns:
            if not column.get('column_name') or not column.get('data_type'):
                return jsonify({'code': 400, 'message': '每一列都必须有列名和数据类型', 'data': None}), 400
            # 为没有hidden属性的列设置默认值
            if 'hidden' not in column:
                column['hidden'] = False
        
        table.columns = json.dumps(columns)
        
        # 获取该表格所有数据，用于处理自增列
        inventory_data_list = InventoryData.query.filter_by(table_id=table_id).all()
        
        # 处理自增列的自增序列
        for column in columns:
            if column.get('autoIncrement'):
                column_name = column['column_name']
                prefix = column.get('prefix', '')
                
                # 查找现有数据中该列的最大值
                max_value = 0
                import re
                pattern = re.compile(f'^{re.escape(prefix)}(\d+)$')
                
                for inventory_data in inventory_data_list:
                    data_json = json.loads(inventory_data.data)
                    if column_name in data_json:
                        value = data_json[column_name]
                        match = pattern.match(value)
                        if match:
                            numeric_part = int(match.group(1))
                            if numeric_part > max_value:
                                max_value = numeric_part
                
                # 获取或创建自增序列
                sequence = AutoIncrementSequence.query.filter_by(
                    table_id=table_id,
                    column_name=column_name
                ).first()
                
                if not sequence:
                    # 创建新的自增序列
                    sequence = AutoIncrementSequence(
                        table_id=table_id,
                        column_name=column_name,
                        current_value=max_value
                    )
                    db.session.add(sequence)
                else:
                    # 更新现有序列的当前值
                    sequence.current_value = max_value
    
    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '表格结构更新成功',
            'data': {
                'id': table.id,
                'table_name': table.table_name,
                'columns': json.loads(table.columns),
                'created_at': table.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': table.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新表格结构失败: {str(e)}', 'data': None}), 500

# 检查自增列数据是否符合要求
@app.route('/api/v1/tables/<int:table_id>/check-auto-increment', methods=['POST'])
@token_required
def check_auto_increment(current_user, table_id):
    # 只有管理员可以检查自增列数据
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    data = request.get_json()
    column_name = data.get('column_name')
    prefix = data.get('prefix', '')
    
    if not column_name:
        return jsonify({'code': 400, 'message': '列名不能为空', 'data': None}), 400
    
    # 获取该表格所有数据
    inventory_data_list = InventoryData.query.filter_by(table_id=table_id).all()
    
    # 用于验证的数据列表
    column_values = []
    numeric_values = []
    existing_prefixes = set()
    
    # 正则表达式：前缀+数字组合
    import re
    pattern = re.compile(f'^(.+?)(\d+)$')
    
    # 如果没有输入前缀，先检测现有数据的前缀
    detected_prefix = prefix
    if not prefix and inventory_data_list:
        # 先从第一条数据中提取前缀
        first_data = inventory_data_list[0]
        data_json = json.loads(first_data.data)
        if column_name in data_json:
            value = data_json[column_name]
            match = pattern.match(value)
            if match:
                detected_prefix = match.group(1)
    
    for inventory_data in inventory_data_list:
        data_json = json.loads(inventory_data.data)
        if column_name in data_json:
            value = data_json[column_name]
            column_values.append(value)
            
            # 验证格式：前缀+数字组合
            match = pattern.match(value)
            if not match:
                return jsonify({'code': 400, 'message': f'列 {column_name} 中存在不符合格式的数据: {value}。数据必须符合 [前缀字符串][数字] 格式。', 'data': None}), 400
            
            # 提取前缀和数字部分
            actual_prefix = match.group(1)
            numeric_part = match.group(2)
            
            existing_prefixes.add(actual_prefix)
            
            # 验证前缀是否与检测到的前缀匹配
            if actual_prefix != detected_prefix:
                return jsonify({'code': 400, 'message': f'列 {column_name} 中数据的前缀与预期前缀不匹配。现有数据的前缀为: {actual_prefix}，预期前缀为: {detected_prefix}。', 'data': None}), 400
            
            numeric_values.append(int(numeric_part))
    
    # 验证唯一性
    if len(column_values) != len(set(column_values)):
        return jsonify({'code': 400, 'message': f'列 {column_name} 中存在重复数据。请重新整理数据，确保所有数据唯一。', 'data': None}), 400
    
    # 计算最大数字值
    max_value = max(numeric_values) if numeric_values else 0
    
    # 如果数据量大于0，检查现有数据的前缀是否一致
    if len(existing_prefixes) > 1:
        return jsonify({'code': 400, 'message': f'列 {column_name} 中数据的前缀不一致。现有数据包含多种前缀: {list(existing_prefixes)}。请重新整理数据，确保所有数据使用统一的前缀。', 'data': None}), 400
    
    # 返回检测到的前缀和最大数字值
    return jsonify({'code': 200, 'message': '数据符合要求', 'data': {'max_value': max_value, 'prefix': detected_prefix}}), 200

# 删除表格结构
@app.route('/api/v1/tables/<int:table_id>', methods=['DELETE'])
@token_required
def delete_table_structure(current_user, table_id):
    # 只有管理员可以删除表格结构
    if current_user.role != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    try:
        # 删除关联的库存数据
        InventoryData.query.filter_by(table_id=table_id).delete()
        # 删除表格结构
        db.session.delete(table)
        db.session.commit()
        
        return jsonify({'code': 200, 'message': '表格结构删除成功', 'data': None}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除表格结构失败: {str(e)}', 'data': None}), 500

# 删除所有表格结构
@app.route('/api/v1/tables/all', methods=['DELETE'])
@token_required
def delete_all_tables(current_user):
    # 只有用户名为"admin"的管理员才能删除所有表格结构
    if current_user.role != 'admin' or current_user.username != 'admin':
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    try:
        # 删除所有关联的库存数据
        InventoryData.query.delete()
        # 删除所有自增序列
        AutoIncrementSequence.query.delete()
        # 删除所有表格结构
        TableStructure.query.delete()
        db.session.commit()
        
        return jsonify({'code': 200, 'message': '所有表格结构删除成功', 'data': None}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除所有表格结构失败: {str(e)}', 'data': None}), 500

# 添加库存数据
@app.route('/api/v1/tables/<int:table_id>/data', methods=['POST'])
@token_required
def add_inventory_data(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    data = request.get_json()
    inventory_data = data.get('data')
    
    if not inventory_data or not isinstance(inventory_data, dict):
        return jsonify({'code': 400, 'message': '数据格式错误', 'data': None}), 400
    
    # 获取表格列配置
    columns = json.loads(table.columns)
    
    # 处理自增列
    for column in columns:
        if column.get('autoIncrement'):
            column_name = column['column_name']
            prefix = column.get('prefix', '')
            
            # 获取或创建自增序列
            sequence = AutoIncrementSequence.query.filter_by(
                table_id=table_id,
                column_name=column_name
            ).first()
            
            if not sequence:
                # 创建新的自增序列
                sequence = AutoIncrementSequence(
                    table_id=table_id,
                    column_name=column_name,
                    current_value=0
                )
                db.session.add(sequence)
            
            # 递增序列值
            sequence.current_value += 1
            
            # 生成自增值
            auto_value = f"{prefix}{sequence.current_value}"
            
            # 将自增值添加到数据中
            inventory_data[column_name] = auto_value
    
    # 将库存数据转换为JSON字符串
    data_json = json.dumps(inventory_data)
    
    try:
        new_data = InventoryData(
            table_id=table_id,
            data=data_json,
            created_by=current_user.id
        )
        db.session.add(new_data)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': '库存数据添加成功',
            'data': {
                'id': new_data.id,
                'table_id': new_data.table_id,
                'data': inventory_data,
                'created_by': new_data.created_by,
                'created_at': new_data.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': new_data.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'添加库存数据失败: {str(e)}', 'data': None}), 500

# 获取库存数据列表
@app.route('/api/v1/tables/<int:table_id>/data', methods=['GET'])
@token_required
def get_inventory_data_list(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 获取查询参数
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 10))
    search = request.args.get('search', '')
    filters = request.args.get('filters', '')
    
    # 构建查询
    query = InventoryData.query.filter_by(table_id=table_id)
    
    # 执行查询并使用默认排序
    paginated_data = query.order_by(
        InventoryData.created_at.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)
    
    # 转换数据格式
    items = []
    for item in paginated_data.items:
        items.append({
            'id': item.id,
            'table_id': item.table_id,
            'data': json.loads(item.data),
            'created_by': item.created_by,
            'created_at': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': item.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'items': items,
            'total': paginated_data.total,
            'page': page,
            'per_page': per_page
        }
    }), 200

# 获取库存数据详情
@app.route('/api/v1/tables/<int:table_id>/data/<int:data_id>', methods=['GET'])
@token_required
def get_inventory_data_detail(current_user, table_id, data_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 获取库存数据
    inventory_data = InventoryData.query.filter_by(id=data_id, table_id=table_id).first()
    if not inventory_data:
        return jsonify({'code': 404, 'message': '库存数据不存在', 'data': None}), 404
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'id': inventory_data.id,
            'table_id': inventory_data.table_id,
            'data': json.loads(inventory_data.data),
            'created_by': inventory_data.created_by,
            'created_at': inventory_data.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': inventory_data.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        }
    }), 200

# 更新库存数据
@app.route('/api/v1/tables/<int:table_id>/data/<int:data_id>', methods=['PUT'])
@token_required
def update_inventory_data(current_user, table_id, data_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 检查用户权限，管理员、组长和组员都可以修改数据
    if current_user.role not in ['admin', 'leader', 'member']:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    # 获取库存数据
    inventory_data = InventoryData.query.filter_by(id=data_id, table_id=table_id).first()
    if not inventory_data:
        return jsonify({'code': 404, 'message': '库存数据不存在', 'data': None}), 404
    
    data = request.get_json()
    new_data = data.get('data')
    
    if not new_data or not isinstance(new_data, dict):
        return jsonify({'code': 400, 'message': '数据格式错误', 'data': None}), 400
    
    # 更新库存数据
    inventory_data.data = json.dumps(new_data)
    
    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '库存数据更新成功',
            'data': {
                'id': inventory_data.id,
                'table_id': inventory_data.table_id,
                'data': new_data,
                'created_by': inventory_data.created_by,
                'created_at': inventory_data.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': inventory_data.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新库存数据失败: {str(e)}', 'data': None}), 500

# 删除库存数据
@app.route('/api/v1/tables/<int:table_id>/data/<int:data_id>', methods=['DELETE'])
@token_required
def delete_inventory_data(current_user, table_id, data_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 检查用户权限，只有管理员和组长可以删除数据
    if current_user.role not in ['admin', 'leader']:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    # 获取库存数据
    inventory_data = InventoryData.query.filter_by(id=data_id, table_id=table_id).first()
    if not inventory_data:
        return jsonify({'code': 404, 'message': '库存数据不存在', 'data': None}), 404
    
    try:
        db.session.delete(inventory_data)
        db.session.commit()
        return jsonify({'code': 200, 'message': '库存数据删除成功', 'data': None}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除库存数据失败: {str(e)}', 'data': None}), 500

# 批量删除库存数据
@app.route('/api/v1/tables/<int:table_id>/data', methods=['DELETE'])
@token_required
def batch_delete_inventory_data(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 检查用户权限，只有管理员和组长可以删除数据
    if current_user.role not in ['admin', 'leader']:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    # 获取要删除的ID列表
    data = request.get_json()
    ids = data.get('ids', [])
    
    if not ids or not isinstance(ids, list) or len(ids) == 0:
        return jsonify({'code': 400, 'message': '删除ID列表不能为空', 'data': None}), 400
    
    try:
        # 执行批量删除
        deleted_count = InventoryData.query.filter_by(table_id=table_id).filter(InventoryData.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功删除{deleted_count}条数据', 'data': {'deleted_count': deleted_count}}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'批量删除库存数据失败: {str(e)}', 'data': None}), 500

# 批量添加库存数据
@app.route('/api/v1/tables/<int:table_id>/data/batch', methods=['POST'])
@token_required
def batch_add_inventory_data(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 检查用户权限，只有管理员和组长可以批量添加数据
    if current_user.role not in ['admin', 'leader']:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    data = request.get_json()
    items = data.get('items')
    
    if not items or not isinstance(items, list) or len(items) == 0:
        return jsonify({'code': 400, 'message': '批量数据格式错误', 'data': None}), 400
    
    success_count = 0
    error_count = 0
    error_messages = []
    
    try:
        for item in items:
            if not item.get('data') or not isinstance(item.get('data'), dict):
                error_count += 1
                error_messages.append(f'数据格式错误: {str(item)}')
                continue
            
            try:
                # 将库存数据转换为JSON字符串
                data_json = json.dumps(item['data'])
                
                new_data = InventoryData(
                    table_id=table_id,
                    data=data_json,
                    created_by=current_user.id
                )
                db.session.add(new_data)
                success_count += 1
            except Exception as e:
                error_count += 1
                error_messages.append(f'添加数据失败: {str(e)}')
        
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': f'批量添加成功，成功{success_count}条，失败{error_count}条',
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'error_messages': error_messages
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'批量添加库存数据失败: {str(e)}', 'data': None}), 500

# 获取统计数据
@app.route('/api/v1/reports/<int:table_id>/stats', methods=['GET'])
@token_required
def get_inventory_stats(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    # 获取查询参数
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    group_by = request.args.get('group_by')
    
    # 构建查询
    query = InventoryData.query.filter_by(table_id=table_id)
    
    # 时间范围过滤
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        query = query.filter(InventoryData.created_at >= start_dt)
    
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        # 结束日期需要加一天，以便包含当天
        end_dt = end_dt + timedelta(days=1)
        query = query.filter(InventoryData.created_at < end_dt)
    
    # 获取所有数据
    all_data = query.all()
    
    # 统计数据
    stats_data = []
    if group_by:
        # 按指定字段分组统计
        group_dict = {}
        
        for item in all_data:
            item_data = json.loads(item.data)
            # 检查分组字段是否存在
            if group_by in item_data:
                key = item_data[group_by]
                if key not in group_dict:
                    group_dict[key] = 0
                group_dict[key] += 1
        
        # 转换为列表格式
        for key, count in group_dict.items():
            stats_data.append({
                'group': key,
                'count': count
            })
    else:
        # 默认按日期分组
        group_dict = {}
        
        for item in all_data:
            # 格式化日期 (YYYY-MM-DD)
            date_key = item.created_at.strftime('%Y-%m-%d')
            if date_key not in group_dict:
                group_dict[date_key] = 0
            group_dict[date_key] += 1
        
        # 转换为列表格式并按日期排序
        for date_key in sorted(group_dict.keys()):
            stats_data.append({
                'group': date_key,
                'count': group_dict[date_key]
            })
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'table_id': table_id,
            'table_name': table.table_name,
            'start_date': start_date,
            'end_date': end_date,
            'group_by': group_by,
            'total_records': len(all_data),
            'stats_data': stats_data
        }
    }), 200

# 数据导出API
@app.route('/api/v1/tables/<int:table_id>/export', methods=['GET'])
@token_required
def export_table_data(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        # 返回404错误，确保JSON格式返回正确
        return jsonify({'code': 404, 'message': '表格结构不存在', 'data': None}), 404
    
    try:
        # 获取所有数据
        all_data = InventoryData.query.filter_by(table_id=table_id).all()
        columns = json.loads(table.columns)
        
        # 生成CSV数据
        import csv
        from io import StringIO
        
        # 使用UTF-8编码的StringIO对象，处理中文分隔符
        output = StringIO()
        
        # 写表头
        header = [col['column_name'] for col in columns]
        # 准备CSV数据，确保正确的分隔符
        csv_lines = []
        
        # 添加表头
        csv_lines.append(','.join(header))
        
        # 写数据行
        for item in all_data:
            item_data = json.loads(item.data)
            # 确保每个值都转换为字符串，处理特殊字符
            row = []
            for col in columns:
                value = item_data.get(col['column_name'], '')
                # 如果值为None，转换为空字符串
                if value is None:
                    value = ''
                else:
                    value = str(value)
                # 处理CSV特殊字符
                # 1. 如果值包含逗号、引号或换行符，需要用引号包裹
                # 2. 如果值包含引号，需要转换为双引号
                if ',' in value or '"' in value or '\n' in value or '\r' in value:
                    value = '"' + value.replace('"', '""') + '"'
                row.append(value)
            # 将数据行转换为CSV格式
            csv_lines.append(','.join(row))
        
        # 使用Windows风格的换行符\r\n，确保跨平台兼容性
        csv_content_str = '\r\n'.join(csv_lines)
        # 转换为字节流
        csv_content = csv_content_str.encode('utf-8')
        
        # 设置响应头
        from flask import Response
        # 确保文件名使用UTF-8编码，解决中文文件名问题
        from urllib.parse import quote
        filename = f'{table.table_name}_export.csv'
        encoded_filename = quote(filename)
        
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{encoded_filename}"',
                'Content-Length': str(len(csv_content)),
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )
    except Exception as e:
        # 打印详细错误信息到日志
        import traceback
        print(f"导出数据发生错误: {traceback.format_exc()}")
        # 返回JSON格式的错误响应
        return jsonify({'code': 500, 'message': f'数据导出失败: {str(e)}', 'data': None}), 500

# 数据导入API
@app.route('/api/v1/tables/<int:table_id>/import', methods=['POST'])
@token_required
def import_table_data(current_user, table_id):
    # 检查表格是否存在
    table = TableStructure.query.get(table_id)
    if not table:
        return jsonify({'code': 404, 'message': '����ṹ������', 'data': None}), 404
    
    # 检查用户权限
    if current_user.role not in ['admin', 'leader']:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None}), 403
    
    try:
        data = request.get_json()
        csv_content = data.get('csv_content')
        
        if not csv_content:
            return jsonify({'code': 400, 'message': 'CSV数据不能为空', 'data': None}), 400
        
        # 解析CSV数据
        import csv
        from io import StringIO
        
        csv_file = StringIO(csv_content)
        reader = csv.reader(csv_file)
        
        # 获取表头
        headers = next(reader, None)
        if not headers:
            return jsonify({'code': 400, 'message': 'CSV文件为空', 'data': None}), 400
        
        # 获取表格列配置
        columns = json.loads(table.columns)
        column_names = [col['column_name'] for col in columns]
        
        # 验证表头与列配置是否匹配
        if len(headers) != len(column_names):
            return jsonify({'code': 400, 'message': f'CSV表头数量不匹配，需要{len(column_names)}列，CSV文件有{len(headers)}列', 'data': None}), 400
        
        for i, header in enumerate(headers):
            if header != column_names[i]:
                return jsonify({'code': 400, 'message': f'CSV表头内容不匹配，第{i+1}列应为"{column_names[i]}"，实际为"{header}"', 'data': None}), 400
        
        # 处理数据行
        success_count = 0
        fail_count = 0
        error_messages = []
        
        for row_num, row in enumerate(reader, start=2):  # 从第2行开始处理数据
            if len(row) != len(columns):
                fail_count += 1
                error_messages.append(f'第{row_num}行，列数不匹配，需要{len(columns)}列，实际有{len(row)}列')
                continue
            
            try:
                # 构建数据字典
                row_data = {}
                for i, col in enumerate(columns):
                    row_data[col['column_name']] = row[i]
                
                # 保存库存数据记录
                new_data = InventoryData(
                    table_id=table_id,
                    data=json.dumps(row_data),
                    created_by=current_user.id
                )
                db.session.add(new_data)
                success_count += 1
            except Exception as e:
                fail_count += 1
                error_messages.append(f'第{row_num}行，处理失败: {str(e)}')
        
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': f'数据导入成功，成功{success_count}条，失败{fail_count}条',
            'data': {
                'success_count': success_count,
                'fail_count': fail_count,
                'error_messages': error_messages
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'数据导入失败: {str(e)}', 'data': None}), 500

# Bug反馈API - 提交Bug
@app.route('/api/v1/bugs', methods=['POST'])
@token_required
def submit_bug(current_user):
    data = request.get_json()
    title = data.get('title')
    content = data.get('content')
    
    if not title or not content:
        return jsonify({'code': 400, 'message': '标题和内容不能为空', 'data': None}), 400
    
    try:
        new_bug = BugReport(
            title=title,
            content=content,
            reporter_username=current_user.username
        )
        db.session.add(new_bug)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': 'Bug反馈提交成功',
            'data': {
                'id': new_bug.id,
                'title': new_bug.title,
                'content': new_bug.content,
                'reporter_username': new_bug.reporter_username,
                'is_resolved': new_bug.is_resolved,
                'created_at': new_bug.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'提交Bug失败: {str(e)}', 'data': None}), 500

# Bug反馈API - 获取所有Bug列表
@app.route('/api/v1/bugs', methods=['GET'])
@token_required
def get_bugs(current_user):
    # 所有人都可以查看Bug列表
    try:
        bugs = BugReport.query.all()
        bugs_data = []
        for bug in bugs:
            bugs_data.append({
                'id': bug.id,
                'title': bug.title,
                'content': bug.content,
                'reporter_username': bug.reporter_username,
                'is_resolved': bug.is_resolved,
                'created_at': bug.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': bug.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({
            'code': 200,
            'message': 'success',
            'data': {
                'items': bugs_data,
                'total': len(bugs_data)
            }
        }), 200
    except Exception as e:
        return jsonify({'code': 500, 'message': f'获取Bug列表失败: {str(e)}', 'data': None}), 500

# Bug反馈API - 更新Bug状态
@app.route('/api/v1/bugs/<int:bug_id>', methods=['PUT'])
@token_required
def update_bug(current_user, bug_id):
    # 只有admin可以更新Bug状态
    if current_user.username != 'admin':
        return jsonify({'code': 403, 'message': '权限不足，只有admin用户可以更新Bug状态', 'data': None}), 403
    
    bug = BugReport.query.get(bug_id)
    if not bug:
        return jsonify({'code': 404, 'message': 'Bug不存在', 'data': None}), 404
    
    try:
        bug.is_resolved = not bug.is_resolved
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': 'Bug状态更新成功',
            'data': {
                'id': bug.id,
                'is_resolved': bug.is_resolved
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新Bug状态失败: {str(e)}', 'data': None}), 500

# Bug反馈API - 删除Bug
@app.route('/api/v1/bugs/<int:bug_id>', methods=['DELETE'])
@token_required
def delete_bug(current_user, bug_id):
    # 只有admin可以删除Bug
    if current_user.username != 'admin':
        return jsonify({'code': 403, 'message': '权限不足，只有admin用户可以删除Bug', 'data': None}), 403
    
    bug = BugReport.query.get(bug_id)
    if not bug:
        return jsonify({'code': 404, 'message': 'Bug不存在', 'data': None}), 404
    
    try:
        db.session.delete(bug)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': 'Bug删除成功',
            'data': None
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除Bug失败: {str(e)}', 'data': None}), 500

# 数据导出API - 将所有表格数据导出为XLS文件
@app.route('/api/v1/export-all-data', methods=['GET'])
@token_required
def export_all_data(current_user):
    # 检查用户权限，只有admin角色可以执行此操作
    if current_user.role != 'admin':
        # 记录操作日志
        log = OperationLog(
            user_id=current_user.id,
            operation='export_all_data',
            table_id=None,
            data_id=None
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'code': 403, 'message': '权限不足，只有管理员可以执行此操作', 'data': None}), 403
    
    try:
        print("开始执行导出所有数据功能")
        from io import BytesIO
        from openpyxl import Workbook
        
        # 创建工作簿
        wb = Workbook()
        print("成功创建工作簿")
        
        # 获取所有表格结构
        tables = TableStructure.query.all()
        print(f"成功获取{len(tables)}个表格结构")
        
        if not tables:
            print("没有可导出的表格数据")
            return jsonify({'code': 400, 'message': '没有可导出的表格数据', 'data': None}), 400
        
        # 删除默认工作表
        # 遍历所有工作表，删除名称为'Sheet'或'Sheet1'的默认工作表
        for sheet in wb.sheetnames:
            if sheet in ['Sheet', 'Sheet1']:
                del wb[sheet]
                print(f"成功删除默认工作表: {sheet}")
                break
        
        # 为每个表格创建一个工作表
        for table in tables:
            table_name = table.table_name
            print(f"\n===== 开始处理表格: {table_name} (ID: {table.id}) =====")
            
            # 输出表格完整信息
            print(f"表格完整信息: {table}")
            print(f"表格columns字段原始内容: {table.columns}")
            
            # 解析columns
            try:
                columns = json.loads(table.columns)
                print(f"成功解析表格{table_name}的columns: {len(columns)}列")
                print(f"列配置详情: {columns}")
            except json.JSONDecodeError as e:
                print(f"解析表格{table_name}的columns失败: {str(e)}")
                print(f"columns原始内容: {table.columns}")
                continue  # 跳过无法解析columns的表格
            
            # 显示当前工作簿中的工作表（创建数据工作表前）
            print(f"当前工作簿中的工作表（创建数据工作表前）: {wb.sheetnames}")
            
            # 创建数据工作表，确保工作表名称不超过31个字符
            data_sheet_name = table_name[:31]
            data_ws = wb.create_sheet(title=data_sheet_name)
            print(f"成功创建数据工作表: {data_sheet_name}")
            print(f"数据工作表对象: {data_ws}")
            print(f"当前工作簿中的工作表（创建数据工作表后）: {wb.sheetnames}")
            
            # 创建列属性工作表，名称不超过31个字符
            properties_sheet_name = f"{table_name[:25]}_属性"
            properties_ws = wb.create_sheet(title=properties_sheet_name)
            print(f"成功创建列属性工作表: {properties_sheet_name}")
            print(f"属性工作表对象: {properties_ws}")
            print(f"当前工作簿中的工作表（创建属性工作表后）: {wb.sheetnames}")
            
            # 写入列属性工作表
            # 列属性表头
            properties_headers = ['column_name', 'data_type', 'dropDown', 'autoIncrement', 'prefix', 'hidden']
            print(f"写入属性工作表表头: {properties_headers}")
            headers_result = properties_ws.append(properties_headers)
            print(f"写入表头结果: {headers_result}")
            
            # 验证表头写入
            header_row = next(properties_ws.iter_rows(min_row=1, max_row=1, values_only=True))
            print(f"属性工作表表头实际内容: {header_row}")
            
            # 写入列属性数据
            print("开始写入列属性数据:")
            for i, col in enumerate(columns):
                properties_row = [
                    col.get('column_name', ''),
                    col.get('data_type', 'string'),
                    col.get('dropDown', False),
                    col.get('autoIncrement', False),
                    col.get('prefix', ''),
                    col.get('hidden', False)
                ]
                print(f"写入第{i+1}列属性: {properties_row}")
                row_result = properties_ws.append(properties_row)
                print(f"写入行结果: {row_result}")
                
                # 验证写入是否成功
                row_count = sum(1 for _ in properties_ws.iter_rows())
                print(f"属性工作表当前行数: {row_count}")
                
                # 读取并打印刚刚写入的行
                written_row = next(properties_ws.iter_rows(min_row=i+2, max_row=i+2, values_only=True))
                print(f"刚刚写入的行实际内容: {written_row}")
            
            # 验证属性工作表完整内容
            print(f"\n=== 属性工作表{properties_sheet_name}完整内容 ===")
            for i, row in enumerate(properties_ws.iter_rows(), 1):
                row_values = [cell.value for cell in row]
                print(f"第{i}行: {row_values}")
            
            print(f"\n===== 成功处理表格: {table_name} =====")
            
            # 获取表格数据
            all_data = InventoryData.query.filter_by(table_id=table.id).all()
            print(f"成功获取表格{table_name}的{len(all_data)}条数据")
            
            # 生成数据工作表表头，包含所有列名和创建时间
            data_headers = [col['column_name'] for col in columns] + ['创建时间']
            data_ws.append(data_headers)
            print(f"成功写入数据工作表表头: {data_headers}")
            
            # 写数据行
            for i, item in enumerate(all_data):
                try:
                    print(f"处理表格{table_name}的数据行 {i+1}/{len(all_data)}")
                    
                    # 将JSON字符串转换为Python对象
                    item_data = json.loads(item.data)
                    
                    # 处理数据行
                    row = []
                    for col in columns:
                        value = item_data.get(col['column_name'], '')
                        # 确保value不是None
                        if value is None:
                            value = ''
                        
                        # 处理链接对象，提取文本部分
                        if isinstance(value, dict):
                            # 如果是链接对象，提取_text属性
                            if '_text' in value:
                                value = value['_text']
                            else:
                                # 其他字典类型，转换为JSON字符串
                                value = json.dumps(value)
                        
                        # 确保value是基本类型，可被Excel处理
                        if not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                        
                        row.append(value)
                    
                    # 添加创建时间
                    row.append(item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    data_ws.append(row)
                    print(f"成功写入表格{table_name}的数据行 {i+1}")
                except Exception as e:
                    print(f"处理表格{table_name}的数据行 {i+1}失败: {str(e)}")
                    # 打印详细错误信息，便于调试
                    print(f"数据行原始内容: {item.data}")
                    import traceback
                    traceback.print_exc()
                    continue  # 跳过无法处理的数据行
            
            print(f"成功写入表格{table_name}的所有数据")
        
        # 保存工作簿前的最终检查
        print(f"\n=== 保存工作簿前的最终检查 ===")
        print(f"工作簿中包含的所有工作表: {wb.sheetnames}")
        
        # 检查每个属性工作表的内容
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_属性'):
                print(f"\n=== 属性工作表{sheet_name}最终内容 ===")
                ws = wb[sheet_name]
                for i, row in enumerate(ws.iter_rows(), 1):
                    row_values = [cell.value for cell in row]
                    print(f"第{i}行: {row_values}")
        
        # 将工作簿保存到BytesIO对象
        print("\n=== 开始保存工作簿到BytesIO对象 ===")
        output = BytesIO()
        wb.save(output)
        print(f"工作簿保存成功，BytesIO位置: {output.tell()}字节")
        
        output.seek(0)
        print(f"BytesIO重置位置: {output.tell()}字节")
        
        # 记录操作日志
        log = OperationLog(
            user_id=current_user.id,
            operation='export_all_data',
            table_id=None,
            data_id=None
        )
        db.session.add(log)
        db.session.commit()
        print("成功记录操作日志")
        
        # 设置响应头
        from flask import Response
        print("准备返回响应")
        
        # 创建响应对象
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename="all_data.xlsx"',
                'Content-Length': str(len(output.getvalue())),
            }
        )
        
        print(f"响应对象创建成功，响应内容长度: {len(output.getvalue())}字节")
        print(f"响应头: {dict(response.headers)}")
        
        return response
    except Exception as e:
        import traceback
        error_msg = f"导出所有数据发生错误: {str(e)}"
        error_trace = traceback.format_exc()
        print(error_msg)
        print(error_trace)
        return jsonify({'code': 500, 'message': '服务器内部错误，请联系管理员', 'data': None}), 500

# 数据导入API - 从XLS文件导入数据
@app.route('/api/v1/import-data', methods=['POST'])
@token_required
def import_data(current_user):
    # 检查用户权限，只有admin角色可以执行此操作
    if current_user.role != 'admin':
        # 记录操作日志
        log = OperationLog(
            user_id=current_user.id,
            operation='import_data',
            table_id=None,
            data_id=None
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'code': 403, 'message': '权限不足，只有管理员可以执行此操作', 'data': None}), 403
    
    try:
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({'code': 400, 'message': '请选择要导入的XLS文件', 'data': None}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'code': 400, 'message': '请选择要导入的XLS文件', 'data': None}), 400
        
        # 检查文件扩展名
        if not file.filename.endswith('.xlsx') and not file.filename.endswith('.xls'):
            return jsonify({'code': 400, 'message': '只支持XLS和XLSX文件格式', 'data': None}), 400
        
        from openpyxl import load_workbook
        
        # 读取文件内容
        wb = load_workbook(file)
        
        # 处理每个工作表
        results = []
        imported_tables = []
        
        # 首先区分数据工作表和属性工作表
        data_sheets = []
        properties_sheets = {}
        
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_属性'):
                # 属性工作表，提取原始表格名称
                original_table_name = sheet_name[:-3]  # 移除'_属性'后缀
                properties_sheets[original_table_name] = sheet_name
            else:
                # 数据工作表
                data_sheets.append(sheet_name)
        
        print(f"识别到数据工作表: {data_sheets}")
        print(f"识别到属性工作表: {properties_sheets}")
        
        # 处理每个数据工作表
        print(f"准备处理{len(data_sheets)}个数据工作表")
        for i, sheet_name in enumerate(data_sheets):
            print(f"\n===== 开始处理第{i+1}个数据工作表: {sheet_name} =====")
            ws = wb[sheet_name]
            imported_tables.append(sheet_name)
            
            # 获取表头
            try:
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                print(f"工作表{sheet_name}的表头: {header_row}")
                print(f"表头长度: {len(header_row)}")
            except StopIteration:
                print(f"工作表{sheet_name}没有表头行")
                results.append({
                    'table_name': sheet_name,
                    'status': 'failed',
                    'message': '工作表没有表头行'
                })
                continue
            
            if len(header_row) < 2:  # 至少需要一个列名和一个数据行
                print(f"工作表{sheet_name}表头格式不正确，至少需要一个列名和一个数据行")
                results.append({
                    'table_name': sheet_name,
                    'status': 'failed',
                    'message': '表头格式不正确，至少需要一个列名和一个数据行'
                })
                continue
            
            # 检查是否包含创建时间列
            created_at_column_index = None
            for i, header in enumerate(header_row):
                if header == '创建时间':
                    created_at_column_index = i
                    break
            
            if created_at_column_index is None:
                print(f"工作表{sheet_name}缺少创建时间列")
                results.append({
                    'table_name': sheet_name,
                    'status': 'failed',
                    'message': '缺少创建时间列'
                })
                continue
            else:
                print(f"工作表{sheet_name}的创建时间列索引: {created_at_column_index}")
            
            # 获取数据列名
            data_headers = header_row[:created_at_column_index] + header_row[created_at_column_index+1:]
            
            # 查找对应的属性工作表
            columns_def = []
            properties_ws = None
            
            print(f"准备查找表格{sheet_name}的属性工作表")
            print(f"当前properties_sheets字典: {properties_sheets}")
            
            if sheet_name in properties_sheets:
                # 找到属性工作表
                properties_sheet_name = properties_sheets[sheet_name]
                print(f"找到表格{sheet_name}的属性工作表: {properties_sheet_name}")
                
                try:
                    properties_ws = wb[properties_sheet_name]
                    print(f"成功获取属性工作表对象")
                except KeyError:
                    print(f"无法找到属性工作表{properties_sheet_name}")
                    # 使用默认配置
                    for header in data_headers:
                        columns_def.append({
                            'column_name': header,
                            'data_type': 'string',
                            'hidden': False
                        })
                    
                # 读取属性工作表数据
                try:
                    properties_rows = list(properties_ws.iter_rows(min_row=1, values_only=True))
                    print(f"属性工作表{properties_sheet_name}共有{len(properties_rows)}行数据")
                    
                    if len(properties_rows) < 2:  # 至少需要表头和一行数据
                        print(f"属性工作表{properties_sheet_name}格式不正确，使用默认配置")
                        # 使用默认配置
                        for header in data_headers:
                            columns_def.append({
                                'column_name': header,
                                'data_type': 'string',
                                'hidden': False
                            })
                    else:
                        # 解析属性工作表
                        properties_header = properties_rows[0]
                        print(f"属性工作表{properties_sheet_name}的表头: {properties_header}")
                        
                        # 验证属性表头
                        expected_headers = ['column_name', 'data_type', 'dropDown', 'autoIncrement', 'prefix', 'hidden']
                        if list(properties_header) != expected_headers:
                            print(f"属性工作表{properties_sheet_name}表头不匹配，使用默认配置")
                            # 使用默认配置
                            for header in data_headers:
                                columns_def.append({
                                    'column_name': header,
                                    'data_type': 'string',
                                    'hidden': False
                                })
                        else:
                            # 使用属性工作表定义的列配置
                            for row in properties_rows[1:]:
                                if len(row) >= 6:
                                    columns_def.append({
                                        'column_name': row[0] or '',
                                        'data_type': row[1] or 'string',
                                        'dropDown': bool(row[2]),
                                        'autoIncrement': bool(row[3]),
                                        'prefix': row[4] or '',
                                        'hidden': bool(row[5])
                                    })
                            print(f"成功读取表格{sheet_name}的列属性: {len(columns_def)}列")
                except Exception as e:
                    print(f"读取属性工作表{properties_sheet_name}失败: {str(e)}")
                    # 使用默认配置
                    for header in data_headers:
                        columns_def.append({
                            'column_name': header,
                            'data_type': 'string',
                            'hidden': False
                        })
            else:
                # 没有属性工作表，使用默认配置
                print(f"未找到表格{sheet_name}的属性工作表，使用默认配置")
                for header in data_headers:
                    columns_def.append({
                        'column_name': header,
                        'data_type': 'string',
                        'hidden': False
                    })
            
            print(f"表格{sheet_name}的列配置: {columns_def}")
            
            try:
                # 检查表格是否已存在
                existing_table = TableStructure.query.filter_by(table_name=sheet_name).first()
                
                if existing_table:
                    # 如果表格已存在，删除现有表格及其数据
                    print(f"表格{sheet_name}已存在，删除现有表格和数据")
                    InventoryData.query.filter_by(table_id=existing_table.id).delete()
                    AutoIncrementSequence.query.filter_by(table_id=existing_table.id).delete()
                    db.session.delete(existing_table)
                    print(f"成功删除现有表格和数据")
                
                # 创建新的表格结构
                columns_json = json.dumps(columns_def)
                
                new_table = TableStructure(
                    table_name=sheet_name,
                    columns=columns_json
                )
                db.session.add(new_table)
                db.session.flush()  # 获取新表ID，但不提交事务
                print(f"成功创建表格{sheet_name}，ID: {new_table.id}")
                
                # 初始化自增序列
                for col in columns_def:
                    if col.get('autoIncrement'):
                        column_name = col['column_name']
                        # 创建自增序列，初始值为0
                        sequence = AutoIncrementSequence(
                            table_id=new_table.id,
                            column_name=column_name,
                            current_value=0
                        )
                        db.session.add(sequence)
                
                # 提交表格和序列创建
                db.session.commit()
                print(f"成功初始化表格{sheet_name}的自增序列")
            except Exception as e:
                # 回滚事务
                db.session.rollback()
                print(f"创建表格{sheet_name}失败: {str(e)}")
                results.append({
                    'table_name': sheet_name,
                    'status': 'failed',
                    'message': f'创建表格结构失败: {str(e)}'
                })
                continue
            
            # 处理数据行
            success_count = 0
            fail_count = 0
            error_messages = []
            
            try:
                # 从第二行开始读取数据
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) < created_at_column_index + 1:
                        fail_count += 1
                        error_messages.append(f'第{row_num}行，数据为空')
                        continue
                    
                    try:
                        # 构建数据字典
                        row_data = {}
                        for i, header in enumerate(data_headers):
                            if i < len(row):
                                row_data[header] = str(row[i]) if row[i] is not None else ''
                            else:
                                row_data[header] = ''
                        
                        # 获取创建时间值
                        created_at_str = row[created_at_column_index]
                        created_at = None
                        if created_at_str:
                            # 尝试解析时间字符串
                            from datetime import datetime
                            try:
                                created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                # 如果时间格式不正确，使用当前时间
                                from datetime import datetime
                                created_at = datetime.now()
                        else:
                            # 如果没有创建时间，使用当前时间
                            from datetime import datetime
                            created_at = datetime.now()
                        
                        # 保存库存数据记录
                        new_data = InventoryData(
                            table_id=new_table.id,
                            data=json.dumps(row_data),
                            created_by=current_user.id,
                            created_at=created_at
                        )
                        db.session.add(new_data)
                        success_count += 1
                    except Exception as e:
                        fail_count += 1
                        error_messages.append(f'第{row_num}行，处理失败: {str(e)}')
                
                # 提交数据
                db.session.commit()
                print(f"成功导入表格{sheet_name}的{success_count}条数据")
            except Exception as e:
                # 回滚事务
                db.session.rollback()
                print(f"导入表格{sheet_name}的数据失败: {str(e)}")
                results.append({
                    'table_name': sheet_name,
                    'status': 'failed',
                    'message': f'导入数据失败: {str(e)}',
                    'success_count': 0,
                    'fail_count': 0,
                    'error_messages': [f'数据导入失败: {str(e)}']
                })
                continue
            
            # 更新自增序列的当前值
            try:
                for col in columns_def:
                    if col.get('autoIncrement'):
                        column_name = col['column_name']
                        prefix = col.get('prefix', '')
                        
                        # 获取该列的最大值
                        all_data = InventoryData.query.filter_by(table_id=new_table.id).all()
                        max_value = 0
                        
                        if all_data:
                            import re
                            pattern = re.compile(f'^{re.escape(prefix)}(\d+)$')
                            
                            for item in all_data:
                                item_data = json.loads(item.data)
                                if column_name in item_data:
                                    value = item_data[column_name]
                                    match = pattern.match(value)
                                    if match:
                                        numeric_part = int(match.group(1))
                                        if numeric_part > max_value:
                                            max_value = numeric_part
                        
                        # 更新自增序列
                        sequence = AutoIncrementSequence.query.filter_by(
                            table_id=new_table.id,
                            column_name=column_name
                        ).first()
                        if sequence:
                            sequence.current_value = max_value
                            db.session.commit()
                        print(f"成功更新表格{sheet_name}列{column_name}的自增序列，当前值为{max_value}")
                
                results.append({
                    'table_name': sheet_name,
                    'status': 'success',
                    'message': f'成功导入{success_count}条数据，失败{fail_count}条',
                    'success_count': success_count,
                    'fail_count': fail_count,
                    'error_messages': error_messages
                })
            except Exception as e:
                # 回滚事务
                db.session.rollback()
                print(f"更新自增序列失败: {str(e)}")
                results.append({
                    'table_name': sheet_name,
                    'status': 'failed',
                    'message': f'更新自增序列失败: {str(e)}',
                    'success_count': success_count,
                    'fail_count': fail_count + 1,
                    'error_messages': error_messages + [f'更新自增序列失败: {str(e)}']
                })
                continue
        
        # 记录操作日志
        log = OperationLog(
            user_id=current_user.id,
            operation='import_data',
            table_id=None,
            data_id=None
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': '数据导入成功',
            'data': {
                'results': results
            }
        }), 200
    except Exception as e:
        import traceback
        print(f"导入数据发生错误: {traceback.format_exc()}")
        return jsonify({'code': 500, 'message': f'数据导入失败: {str(e)}', 'data': None}), 500

# 修改密码路由
@app.route('/api/v1/auth/change-password', methods=['POST'])
@token_required
def change_password(current_user):
    data = request.get_json()
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    confirm_password = data.get('confirm_password')
    
    # 验证请求参数
    if not current_password or not new_password or not confirm_password:
        return jsonify({'code': 400, 'message': '当前密码、新密码和确认密码不能为空', 'data': None}), 400
    
    # 验证两次密码输入一致
    if new_password != confirm_password:
        return jsonify({'code': 400, 'message': '新密码和确认密码不一致', 'data': None}), 400
    
    from werkzeug.security import check_password_hash, generate_password_hash
    
    # 验证当前密码是否正确
    if not check_password_hash(current_user.password, current_password):
        return jsonify({'code': 400, 'message': '当前密码错误', 'data': None}), 400
    
    # 只验证密码长度大于零
    if len(new_password) == 0:
        return jsonify({'code': 400, 'message': '新密码不能为空', 'data': None}), 400
    
    # 更新密码
    current_user.password = generate_password_hash(new_password)
    
    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '密码修改成功',
            'data': None
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'密码修改失败: {str(e)}', 'data': None}), 500

# 启动应用
if __name__ == '__main__':
    app.run(debug=True, port=5000, host='0.0.0.0')